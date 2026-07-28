#!/usr/bin/env python3
"""Parse MLB_2026_HOMERUNS.xlsx day-sheets into data/homeruns-2026.json with code tags.
Rerun whenever new weeks are added to the workbook. Categories = regex map below."""
import openpyxl, re, json, sys
from collections import Counter
CATS = {
 "birthday_span": r"\bbday|birthday\b",
 "team_value": r"(Royals|Yankees|Dodgers|Cubs|Mets|Braves|Phillies|Marlins|Nationals|Pirates|Reds|Brewers|Cardinals|Rockies|Giants|Padres|Diamondbacks|Astros|Rangers|Mariners|Athletics|Angels|Twins|Guardians|Tigers|White Sox|Red Sox|Orioles|Rays|Blue Jays)\s*\(?\d",
 "city_value": r"(Kansas City|New York|Chicago|Miami|Boston|Seattle|Houston|Atlanta|Denver|Phoenix|Sacramento|Milwaukee|Minneapolis|Pittsburgh|Philadelphia|Cincinnati|Cleveland|Detroit|Baltimore|Tampa|Arlington|Anaheim|San Diego|San Francisco|St\.? Louis|Washington|Los Angeles|Toronto)\s*\(?\d",
 "ordinal_hr_milestone": r"\d+(st|nd|rd|th)\s*(career\s*)?(HR|hr|homerun|home run)",
 "prime_index": r"prime", "composite_index": r"composite|comp\b",
 "day_name_value": r"(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\s*\(?\d",
 "dn_match": r"\bDN\b|\d+\s*DN", "dliy_doy": r"DLIY|DOY|day of year|days left",
 "days_since_last_hr": r"(since|ago).*(hr|HR|homer)|d\s+since", "debut_anniv": r"debut",
 "beast_lucifer_satan": r"Beast|Lucifer|Satan\b", "pope": r"Pope",
 "jesuit_mason": r"Jesuit|Mason|Freemason", "jersey_number": r"#\d+|jersey",
 "name_gematria": r"(Homerun|Homer|Dinger|Moonshot)\s*\(?=?\s*\d|\w+\s+\w+\s*=\s*\d",
 "planet_value": r"Mars|Mercury|Venus|Jupiter|Saturn\b|Moon\s*\(?\d|Sun\s*\(?\d",
 "date_written": r"\d+/\d+|Twenty|July|June|May|April|August",
 "stadium_park": r"Park\s*\(?\d|Field\s*\(?\d|Stadium\s*\(?\d",
 "rbi_pa_ab_milestone": r"\d+(st|nd|rd|th)\s*(RBI|PA|AB|hit|run|XBH|xbh)",
 "game_count": r"\d+(st|nd|rd|th)\s*(career\s*)?game",
 "h2h_vs_split": r"h2h|vs\s+(AL|NL|LHP|RHP|lhp|rhp)",
}
def main(xlsx, out):
    wb = openpyxl.load_workbook(xlsx, data_only=True); rows=[]
    for sn in wb.sheetnames:
        for r in wb[sn].iter_rows(min_row=2, values_only=True):
            if not r or not r[4]: continue
            note = str(r[11]).strip() if r[11] else None
            codes = [c for c,p in CATS.items() if note and re.search(p, note, re.I)]
            rows.append({"date": r[0].date().isoformat() if hasattr(r[0],'date') else str(r[0]),
                "weekday": sn, "dn": r[1], "lp": r[2], "player": str(r[4]).strip(),
                "jersey": r[5], "zodiac": str(r[6]).strip() if r[6] else None,
                "moon_sign": r[7], "moon_phase": r[8], "hr": r[9], "odds": r[10],
                "note": note, "codes": codes})
    json.dump(rows, open(out,'w'), indent=0)
    t=Counter(c for x in rows for c in x['codes'])
    print(f"{len(rows)} rows -> {out}"); [print(f"  {v:5d} {c}") for c,v in t.most_common()]
if __name__=='__main__':
    main(sys.argv[1] if len(sys.argv)>1 else 'MLB_2026_HOMERUNS.xlsx',
         sys.argv[2] if len(sys.argv)>2 else 'data/homeruns-2026.json')
